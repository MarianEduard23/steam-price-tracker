"""
CSFLOAT ITEMS PRICE CALCULATOR - CU GRAFICE HTML INTERACTIVE
Fiecare rulare = un punct pe grafic, per item.
Graficele sunt HTML interactive (zoom, pan, filtre temporale).
Ia cel mai mic preț "Buy Now" de pe CSFloat (ignoră bids).
"""

import requests
import time
from datetime import datetime, timedelta
import json
import os
import html
import pickle

try:
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload
    GDRIVE_AVAILABLE = True
except ImportError:
    GDRIVE_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Librăria 'openpyxl' nu este instalată!")
    print("   pip install openpyxl")
    exit()

# ── Constante ──────────────────────────────────────────────
API_KEY             = os.environ.get("CSFLOAT_API_KEY", "PUNE_API_KEY_AICI")
CSFLOAT_TAX         = 0.02          # 2% taxa CSFloat
MAX_RETRIES         = 3
DELAY_BETWEEN_REQS  = 2             # secunde între request-uri
HISTORY_FILE        = "csfloat_price_history.json"
CHARTS_DIR          = "csfloat_charts"
GDRIVE_FOLDER_NAME  = "CSFloatPriceTracker"
TOKEN_FILE          = "gdrive_token.pickle"
CREDENTIALS_FILE    = "credentials.json"
SCOPES              = ["https://www.googleapis.com/auth/drive.file"]

BASE_URL            = "https://csfloat.com/api/v1"
HEADERS_API         = {
    "Authorization": API_KEY,
    "Content-Type": "application/json",
}

# ── Google Drive ─────────────────────────────────────────────
def get_gdrive_service():
    if not GDRIVE_AVAILABLE:
        return None
    creds = None

    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, "rb") as f:
            creds = pickle.load(f)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            with open(TOKEN_FILE, "wb") as f:
                pickle.dump(creds, f)
        else:
            gdrive_creds_env = os.environ.get("GDRIVE_CREDENTIALS")
            if gdrive_creds_env:
                import tempfile
                with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as tmp:
                    tmp.write(gdrive_creds_env)
                    tmp_path = tmp.name
                flow = InstalledAppFlow.from_client_secrets_file(tmp_path, SCOPES)
                os.unlink(tmp_path)
            elif os.path.exists(CREDENTIALS_FILE):
                flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            else:
                print("❌ credentials.json lipsește și GDRIVE_CREDENTIALS nu e setat!")
                return None

            flow.redirect_uri = "urn:ietf:wg:oauth:2.0:oob"
            auth_url, _ = flow.authorization_url(prompt="consent")
            print("\n🔗 Deschide acest link în browser:\n")
            print(auth_url)
            print()
            code = input("📋 Paste codul primit de la Google aici: ").strip()
            flow.fetch_token(code=code)
            creds = flow.credentials
            with open(TOKEN_FILE, "wb") as f:
                pickle.dump(creds, f)

    return build("drive", "v3", credentials=creds)


def get_or_create_folder(service, name, parent_id=None):
    q = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent_id:
        q += f" and '{parent_id}' in parents"
    res = service.files().list(q=q, fields="files(id)").execute()
    files = res.get("files", [])
    if files:
        return files[0]["id"]
    meta = {"name": name, "mimeType": "application/vnd.google-apps.folder"}
    if parent_id:
        meta["parents"] = [parent_id]
    return service.files().create(body=meta, fields="id").execute().get("id")


def upload_file(service, folder_id, filepath, mime="application/octet-stream"):
    filename = os.path.basename(filepath)
    q = f"name='{filename}' and '{folder_id}' in parents and trashed=false"
    res = service.files().list(q=q, fields="files(id)").execute()
    files = res.get("files", [])
    media = MediaFileUpload(filepath, mimetype=mime, resumable=True)
    if files:
        service.files().update(fileId=files[0]["id"], media_body=media).execute()
        return files[0]["id"]
    meta = {"name": filename, "parents": [folder_id]}
    return service.files().create(body=meta, media_body=media, fields="id").execute().get("id")


def upload_all_to_gdrive(excel_path, charts_dir):
    print("\n☁️  Upload Google Drive...")
    service = get_gdrive_service()
    if not service:
        print("   ⚠️  Skipping upload.")
        return None
    folder_id = get_or_create_folder(service, GDRIVE_FOLDER_NAME)
    if os.path.exists(excel_path):
        upload_file(service, folder_id, excel_path,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        print(f"   ✅ {os.path.basename(excel_path)}")
    if os.path.exists(charts_dir):
        cf_id = get_or_create_folder(service, "csfloat_charts", folder_id)
        html_files = [f for f in os.listdir(charts_dir) if f.endswith(".html")]
        for hf in html_files:
            upload_file(service, cf_id, os.path.join(charts_dir, hf), "text/html")
        print(f"   ✅ {len(html_files)} grafice HTML")
    link = f"https://drive.google.com/drive/folders/{folder_id}"
    print(f"   🔗 {link}")
    with open("gdrive_link.txt", "w") as f:
        f.write(f"{link}\nActualizat: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
    return link


# ── Rata de schimb USD → EUR ────────────────────────────────
def get_usd_to_eur() -> float:
    """Ia rata de schimb USD->EUR live de la api.frankfurter.app."""
    try:
        r = requests.get("https://api.frankfurter.app/latest?from=USD&to=EUR", timeout=10)
        if r.status_code == 200:
            rate = r.json()["rates"]["EUR"]
            print(f"💱 Rata USD->EUR: 1 USD = {rate:.4f} EUR")
            return rate
    except Exception as e:
        print(f"⚠️ Nu s-a putut obtine rata de schimb: {e}")
    print("⚠️ Folosesc rata fallback: 1 USD = 0.92 EUR")
    return 0.92

USD_TO_EUR = 0.92  # actualizat in main() cu rata live

# ── CSFloat API ─────────────────────────────────────────────
def get_csfloat_price(item_name: str, retry: int = 0) -> float | None:
    """
    Returnează cel mai mic preț 'Buy Now' pentru un item pe CSFloat.
    Ignoră complet bid-urile — caută doar listings de tip buy_now.
    Prețul e returnat în EUR (conversie live USD->EUR).
    """
    try:
        params = {
            "market_hash_name": item_name,
            "sort_by": "lowest_price",
            "order": "asc",
            "limit": 5,
            "type": "buy_now",   # filtru corect pentru GET: doar Buy Now, fără bids
        }
        r = requests.get(
            f"{BASE_URL}/listings",
            headers=HEADERS_API,
            params=params,
            timeout=15,
        )

        if r.status_code == 429:
            if retry < MAX_RETRIES:
                wait = 15 * (retry + 1)
                print(f"   ⏳ Rate limit! Aștept {wait}s...")
                time.sleep(wait)
                return get_csfloat_price(item_name, retry + 1)
            print("   ⚠️ Rate limit depășit, skip.")
            return None

        if r.status_code == 401:
            print("   ❌ API key invalid! Verifică constanta API_KEY.")
            return None

        if r.status_code != 200:
            print(f"   ⚠️ HTTP {r.status_code} — răspuns: {r.text[:200]}")
            return None

        data = r.json()

        # API-ul poate returna fie {"data": [...]} fie direct o listă
        listings = data.get("data", data) if isinstance(data, dict) else data

        if not listings:
            print("   ⚠️ Indisponibil pe CSFloat Buy Now")
            return None

        # Filtrăm manual orice listing care nu e buy_now (extra siguranță)
        buy_now = [l for l in listings if l.get("type", "buy_now") == "buy_now"]
        if not buy_now:
            buy_now = listings  # dacă câmpul "type" nu există, luăm tot

        # Pretul e in centi USD -> impartim la 100 si convertim in EUR
        lowest = min(buy_now, key=lambda l: l["price"])
        price_usd = lowest["price"] / 100.0
        return round(price_usd * USD_TO_EUR, 2)

    except Exception as e:
        print(f"   ⚠️ Eroare: {e}")
        return None


# ── Istoric JSON ────────────────────────────────────────────
def load_history() -> dict:
    if not os.path.exists(HISTORY_FILE):
        return {}
    try:
        with open(HISTORY_FILE, "r", encoding="utf-8") as f:
            raw = json.load(f)
        out = {}
        for k, v in raw.items():
            if k == "__global_total__":
                if "history" not in v:
                    out[k] = {"history": [{"value": v.get("value", 0), "timestamp": v.get("timestamp")}]}
                else:
                    out[k] = v
            else:
                if "history" not in v:
                    out[k] = {
                        "current": {"total": v.get("total", 0), "price": v.get("price", 0),
                                    "quantity": v.get("quantity", 1), "timestamp": v.get("timestamp")},
                        "history": [{"total": v.get("total", 0), "price": v.get("price", 0),
                                     "timestamp": v.get("timestamp")}]
                    }
                else:
                    out[k] = v
        return out
    except:
        return {}


def save_history(h: dict):
    try:
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(h, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️ Nu s-a putut salva istoricul: {e}")


# ── Căutare preț la dată ────────────────────────────────────
def price_at(history_list: list, days_ago: int, tolerance: int):
    if not history_list:
        return None
    now    = datetime.now()
    target = now - timedelta(days=days_ago)
    lo     = target - timedelta(days=tolerance)
    hi     = target + timedelta(days=tolerance)
    best, best_diff = None, float("inf")
    for e in history_list:
        try:
            dt = datetime.fromisoformat(e["timestamp"])
            if lo <= dt <= hi:
                d = abs((dt - target).total_seconds())
                if d < best_diff:
                    best_diff, best = d, e
        except:
            continue
    return best["total"] if best else None


# ── Generare HTML Chart ─────────────────────────────────────
def generate_html_chart(item_name: str, hist_list: list, chart_path: str):
    """Generează un fișier HTML cu grafic interactiv dark-theme cu Plotly."""
    sorted_hist = sorted(hist_list, key=lambda x: x.get("timestamp", ""))

    timestamps = []
    prices     = []
    for entry in sorted_hist:
        try:
            dt = datetime.fromisoformat(entry["timestamp"])
            timestamps.append(dt.strftime("%Y-%m-%d %H:%M"))
            prices.append(round(entry["total"], 4))
        except:
            continue

    if not timestamps:
        return

    ts_json   = json.dumps(timestamps)
    pr_json   = json.dumps(prices)
    safe_name = html.escape(item_name)

    html_content = f"""<!DOCTYPE html>
<html lang="ro">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{safe_name}</title>
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{
    background: #1b2838;
    color: #c6d4df;
    font-family: 'Motiva Sans', Arial, sans-serif;
    display: flex;
    flex-direction: column;
    height: 100vh;
    overflow: hidden;
  }}
  .header {{
    padding: 14px 20px 8px;
    background: #171a21;
    border-bottom: 1px solid #2a475e;
    flex-shrink: 0;
  }}
  .header h1 {{
    font-size: 15px;
    color: #c6d4df;
    font-weight: 600;
    margin-bottom: 4px;
  }}
  .header .subtitle {{
    font-size: 11px;
    color: #8f98a0;
  }}
  .btn-row {{
    display: flex;
    gap: 6px;
    padding: 10px 20px;
    background: #171a21;
    flex-shrink: 0;
    align-items: center;
    border-bottom: 1px solid #2a475e;
  }}
  .btn-row span {{
    font-size: 11px;
    color: #8f98a0;
    margin-right: 4px;
  }}
  button {{
    background: #2a475e;
    color: #c6d4df;
    border: 1px solid #3d6b8c;
    border-radius: 3px;
    padding: 5px 14px;
    font-size: 12px;
    cursor: pointer;
    transition: background 0.15s, color 0.15s;
  }}
  button:hover {{ background: #3d6b8c; color: #fff; }}
  button.active {{
    background: #e67e22;
    color: #fff;
    border-color: #e67e22;
    font-weight: 700;
  }}
  #chart {{
    flex: 1;
    min-height: 0;
  }}
  .stats-bar {{
    display: flex;
    gap: 20px;
    padding: 8px 20px;
    background: #171a21;
    border-top: 1px solid #2a475e;
    font-size: 11px;
    color: #8f98a0;
    flex-shrink: 0;
  }}
  .stat {{ display: flex; gap: 5px; }}
  .stat-val {{ color: #c6d4df; font-weight: 600; }}
  .stat-val.up {{ color: #e67e22; }}
  .stat-val.down {{ color: #e74c3c; }}
  .csfloat-badge {{
    margin-left: auto;
    font-size: 10px;
    color: #e67e22;
    font-weight: 600;
    letter-spacing: 0.5px;
  }}
</style>
</head>
<body>
<div class="header">
  <h1>📊 {safe_name}</h1>
  <div class="subtitle">CSFloat — Cel mai mic preț Buy Now ($) — fiecare punct = o rulare a scriptului</div>
</div>
<div class="btn-row">
  <span>Interval:</span>
  <button onclick="filterData(3)" id="btn3d">3 Zile</button>
  <button onclick="filterData(7)" id="btn7d">1 Săpt.</button>
  <button onclick="filterData(30)" id="btn30d">1 Lună</button>
  <button onclick="filterData(90)" id="btn90d">3 Luni</button>
  <button onclick="filterData(180)" id="btn180d">6 Luni</button>
  <button onclick="filterData(0)" id="btnAll" class="active">Tot</button>
</div>
<div id="chart"></div>
<div class="stats-bar" id="statsBar">
  <div class="stat">Date: <span class="stat-val" id="statPoints">—</span> puncte</div>
  <div class="stat">Min: <span class="stat-val" id="statMin">—</span></div>
  <div class="stat">Max: <span class="stat-val" id="statMax">—</span></div>
  <div class="stat">Medie: <span class="stat-val" id="statAvg">—</span></div>
  <div class="stat">Curent: <span class="stat-val" id="statCurrent">—</span></div>
  <div class="stat">vs Prima: <span class="stat-val" id="statChange">—</span></div>
  <div class="csfloat-badge">⚡ CSFloat Buy Now</div>
</div>

<script>
const allTimestamps = {ts_json};
const allPrices = {pr_json};
let currentChart = null;

function buildTrace(ts, prices) {{
  return {{
    x: ts,
    y: prices,
    type: 'scatter',
    mode: 'lines+markers',
    name: 'Preț Total ($)',
    line: {{
      color: '#e67e22',
      width: 2,
      shape: 'linear'
    }},
    marker: {{
      color: '#e67e22',
      size: 5,
    }},
    fill: 'tozeroy',
    fillcolor: 'rgba(230,126,34,0.07)',
    hovertemplate: '<b>%{{x}}</b><br>€%{{y:.4f}}<extra></extra>'
  }};
}}

const layout = {{
  paper_bgcolor: '#1b2838',
  plot_bgcolor:  '#1b2838',
  margin: {{ t: 10, r: 20, b: 45, l: 65 }},
  xaxis: {{
    color: '#8f98a0',
    gridcolor: '#2a475e',
    linecolor: '#2a475e',
    tickfont: {{ size: 10, color: '#8f98a0' }},
    showspikes: true,
    spikecolor: '#e67e22',
    spikethickness: 1,
    spikedash: 'dot',
    spikemode: 'across',
  }},
  yaxis: {{
    color: '#8f98a0',
    gridcolor: '#2a475e',
    linecolor: '#2a475e',
    tickfont: {{ size: 10, color: '#8f98a0' }},
    tickprefix: '€',
    showspikes: true,
    spikecolor: '#e67e22',
    spikethickness: 1,
    spikedash: 'dot',
  }},
  hoverlabel: {{
    bgcolor: '#2a475e',
    bordercolor: '#e67e22',
    font: {{ color: '#c6d4df', size: 12 }}
  }},
  dragmode: 'zoom',
  selectdirection: 'h',
}};

const config = {{
  responsive: true,
  displaylogo: false,
  modeBarButtonsToRemove: ['lasso2d', 'select2d', 'autoScale2d'],
  scrollZoom: true,
}};

function updateStats(prices) {{
  if (!prices.length) return;
  const min   = Math.min(...prices);
  const max   = Math.max(...prices);
  const avg   = prices.reduce((a, b) => a + b, 0) / prices.length;
  const cur   = prices[prices.length - 1];
  const first = prices[0];
  const chg   = ((cur - first) / first * 100);
  document.getElementById('statPoints').textContent  = prices.length;
  document.getElementById('statMin').textContent     = '€' + min.toFixed(4);
  document.getElementById('statMax').textContent     = '€' + max.toFixed(4);
  document.getElementById('statAvg').textContent     = '€' + avg.toFixed(4);
  document.getElementById('statCurrent').textContent = '€' + cur.toFixed(4);
  const chgEl = document.getElementById('statChange');
  chgEl.textContent = (chg >= 0 ? '+' : '') + chg.toFixed(2) + '%';
  chgEl.className   = 'stat-val ' + (chg > 0 ? 'up' : chg < 0 ? 'down' : '');
}}

function filterData(days) {{
  ['btn3d','btn7d','btn30d','btn90d','btn180d','btnAll'].forEach(id => {{
    document.getElementById(id).classList.remove('active');
  }});
  const btnMap = {{3:'btn3d',7:'btn7d',30:'btn30d',90:'btn90d',180:'btn180d',0:'btnAll'}};
  if (btnMap[days]) document.getElementById(btnMap[days]).classList.add('active');

  let ts = allTimestamps;
  let pr = allPrices;

  if (days > 0) {{
    const cutoff  = new Date();
    cutoff.setDate(cutoff.getDate() - days);
    const filtered = allTimestamps.map((t, i) => [t, allPrices[i]])
      .filter(([t]) => new Date(t) >= cutoff);
    ts = filtered.map(([t]) => t);
    pr = filtered.map(([, p]) => p);
  }}

  Plotly.react('chart', [buildTrace(ts, pr)], layout, config);
  updateStats(pr);
}}

filterData(0);
</script>
</body>
</html>"""

    os.makedirs(os.path.dirname(chart_path) if os.path.dirname(chart_path) else ".", exist_ok=True)
    with open(chart_path, "w", encoding="utf-8") as f:
        f.write(html_content)


# ── Stiluri helper ──────────────────────────────────────────
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)
GREEN_FILL   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ALT_FILL     = PatternFill(start_color="F5F8FF", end_color="F5F8FF", fill_type="solid")
HEADER_FILL  = PatternFill(start_color="C0501A", end_color="C0501A", fill_type="solid")  # portocaliu CSFloat
TOTAL_FILL   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TAX_FILL     = PatternFill(start_color="3B2A2A", end_color="3B2A2A", fill_type="solid")
NET_FILL     = PatternFill(start_color="1E3A28", end_color="1E3A28", fill_type="solid")


def cell_font(bold=False, color="000000", size=10, underline=None):
    return Font(name="Arial", bold=bold, color=color, size=size, underline=underline)


# ── Generare Excel ──────────────────────────────────────────
def create_excel(items_list: list, output: str = "csfloat_items.xlsx"):
    os.makedirs(CHARTS_DIR, exist_ok=True)
    wb = Workbook()

    ws       = wb.active
    ws.title = "Iteme CSFloat"
    ws.freeze_panes = "A2"

    HEADERS = [
        "Nume Item", "Cant.", "Preț/Unitate (€)", "Total (€)",
        "vs Ult. Rulare",
        "3 Zile (€)",  "vs 3 Zile",
        "5 Zile (€)",  "vs 5 Zile",
        "7 Zile (€)",  "vs 7 Zile",
        "2 Săpt. (€)", "vs 2 Săpt.",
        "3 Săpt. (€)", "vs 3 Săpt.",
        "30 Zile (€)", "vs 30 Zile",
        "2 Luni (€)",  "vs 2 Luni",
        "3 Luni (€)",  "vs 3 Luni",
        "4 Luni (€)",  "vs 4 Luni",
        "5 Luni (€)",  "vs 5 Luni",
        "6 Luni (€)",  "vs 6 Luni",
        "Actualizat"
    ]
    COL_WIDTHS = [46, 7, 17, 15, 15,
                  14, 12, 14, 12,
                  14, 12, 14, 12,
                  14, 12, 14, 12,
                  14, 12, 14, 12,
                  14, 12, 14, 12,
                  14, 12, 20]

    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill      = HEADER_FILL
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = THIN_BORDER
    ws.row_dimensions[1].height = 36
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Preluare prețuri ────────────────────────────────────
    history   = load_history()
    now_ts    = datetime.now().isoformat()
    items_out = []
    ok = fail = 0

    PERIODS = [
        (3,   1,  "p3d",   "pct3d"),
        (5,   1,  "p5d",   "pct5d"),
        (7,   1,  "p7d",   "pct7d"),
        (14,  1,  "p14d",  "pct14d"),
        (21,  2,  "p21d",  "pct21d"),
        (30,  2,  "p30d",  "pct30d"),
        (60,  4,  "p60d",  "pct60d"),
        (90,  5,  "p90d",  "pct90d"),
        (120, 5,  "p120d", "pct120d"),
        (150, 5,  "p150d", "pct150d"),
        (180, 7,  "p180d", "pct180d"),
    ]

    print("\n🔄 Procesare iteme...\n" + "=" * 60)

    for idx, item in enumerate(items_list, 1):
        name = item["name"]
        qty  = item.get("quantity", 1)
        key  = f"{name}__{qty}"

        print(f"\n[{idx}/{len(items_list)}] {name}")
        price = get_csfloat_price(name)

        if price is not None:
            print(f"   ✅ ${price:.2f}")
            ok += 1
        else:
            print("   ❌ Indisponibil pe CSFloat Buy Now")
            price = 0.0
            fail += 1

        total     = round(price * qty, 4)
        item_hist = history.get(key, {})
        hist_list = item_hist.get("history", [])
        last_total = item_hist.get("current", {}).get("total")

        pct_last = None
        if last_total and last_total > 0 and total > 0:
            pct_last = (total - last_total) / last_total * 100

        pd_vals = {}
        for days, tol, pk, pctk in PERIODS:
            hp = price_at(hist_list, days, tol)
            pd_vals[pk]   = hp
            pd_vals[pctk] = ((total - hp) / hp * 100) if (hp and hp > 0 and total > 0) else None

        hist_list.append({"total": total, "price": price, "timestamp": now_ts})
        history[key] = {
            "current": {"total": total, "price": price, "quantity": qty, "timestamp": now_ts},
            "history": hist_list
        }

        items_out.append({
            "name": name, "price": price, "qty": qty,
            "total": total, "pct_last": pct_last,
            "hist": hist_list,
            **pd_vals,
            "ts": datetime.now().strftime("%Y-%m-%d %H:%M")
        })

        if idx < len(items_list):
            time.sleep(DELAY_BETWEEN_REQS)

    items_out.sort(key=lambda x: x["total"], reverse=True)

    # ── Generare grafice HTML ───────────────────────────────
    print("\n📊 Generare grafice HTML...")
    chart_files = {}
    for d in items_out:
        safe_fn    = "".join(c if c.isalnum() or c in " -_()" else "_" for c in d["name"])[:80]
        chart_path = os.path.join(CHARTS_DIR, f"{safe_fn}.html")
        generate_html_chart(d["name"], d["hist"], chart_path)
        chart_files[d["name"]] = chart_path
        print(f"   📈 {d['name'][:55]}")

    # ── Helper celule ───────────────────────────────────────
    def pct_cell(row, col, val, alt):
        c = ws.cell(row=row, column=col)
        if val is not None:
            c.value     = f"{val:+.2f}%"
            c.alignment = Alignment(horizontal="center")
            if val > 0:
                c.fill = GREEN_FILL
                c.font = cell_font(bold=True, color="006100")
            elif val < 0:
                c.fill = RED_FILL
                c.font = cell_font(bold=True, color="9C0006")
            else:
                c.font = cell_font()
                if alt: c.fill = ALT_FILL
        else:
            c.value     = "N/A"
            c.alignment = Alignment(horizontal="center")
            c.font      = cell_font(color="AAAAAA")
            if alt: c.fill = ALT_FILL
        c.border = THIN_BORDER

    def price_cell(row, col, val, alt):
        c = ws.cell(row=row, column=col)
        if val is not None:
            c.value         = val
            c.number_format = "0.00"
            c.font          = cell_font()
        else:
            c.value     = "N/A"
            c.alignment = Alignment(horizontal="center")
            c.font      = cell_font(color="AAAAAA")
        if alt: c.fill = ALT_FILL
        c.border = THIN_BORDER

    PERIOD_KEYS = [
        ("p3d","pct3d"), ("p5d","pct5d"), ("p7d","pct7d"),
        ("p14d","pct14d"), ("p21d","pct21d"), ("p30d","pct30d"),
        ("p60d","pct60d"), ("p90d","pct90d"), ("p120d","pct120d"),
        ("p150d","pct150d"), ("p180d","pct180d")
    ]

    # ── Scriere rânduri ─────────────────────────────────────
    current_row = 2
    for d in items_out:
        alt  = (current_row % 2 == 0)
        name = d["name"]

        c1 = ws.cell(row=current_row, column=1)
        if name in chart_files:
            # Path relativ fata de Excel: csfloat_charts/NumeItem.html
            rel_path = chart_files[name].replace("\\", "/")
            c1.value     = "📊 " + name
            c1.hyperlink = rel_path
            c1.font      = cell_font(color="C0501A", underline="single")
        else:
            c1.value = name
            c1.font  = cell_font()
        if alt: c1.fill = ALT_FILL
        c1.border = THIN_BORDER

        def plain(col, val, fmt=None):
            c = ws.cell(row=current_row, column=col, value=val)
            c.font   = cell_font()
            c.border = THIN_BORDER
            if fmt: c.number_format = fmt
            if alt: c.fill = ALT_FILL

        plain(2, d["qty"])
        plain(3, d["price"], "0.00")
        plain(4, d["total"], "0.00")
        pct_cell(current_row, 5, d["pct_last"], alt)

        col = 6
        for pk, pctk in PERIOD_KEYS:
            price_cell(current_row, col,     d.get(pk),   alt)
            pct_cell  (current_row, col + 1, d.get(pctk), alt)
            col += 2

        plain(col, d["ts"])
        current_row += 1

    total_sum = sum(d["total"] for d in items_out)
    gh        = history.get("__global_total__", {})
    gh_list   = gh.get("history", [])
    last_g    = gh_list[-1]["value"] if gh_list else None
    pct_g     = ((total_sum - last_g) / last_g * 100) if (last_g and last_g > 0) else None
    gh_list.append({"value": total_sum, "timestamp": now_ts})
    history["__global_total__"] = {"history": gh_list}

    period_totals = {}
    for pk, _ in PERIOD_KEYS:
        vals = [d[pk] for d in items_out if d.get(pk) is not None]
        period_totals[pk] = sum(vals) if vals else None

    def total_header_cell(row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = cell_font(bold=True, color="FFFFFF", size=10)
        c.fill      = TOTAL_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = THIN_BORDER

    def total_value_cell(row, col, value, color="FFFFFF", fmt="0.00"):
        c = ws.cell(row=row, column=col, value=value)
        c.font          = cell_font(bold=True, color=color, size=10)
        c.fill          = TOTAL_FILL
        c.number_format = fmt
        c.alignment     = Alignment(horizontal="center")
        c.border        = THIN_BORDER

    def total_pct_cell(row, col, val):
        c = ws.cell(row=row, column=col)
        c.fill   = TOTAL_FILL
        c.border = THIN_BORDER
        if val is not None:
            c.value     = f"{val:+.2f}%"
            c.alignment = Alignment(horizontal="center")
            c.font      = cell_font(bold=True, color="E67E22" if val > 0 else ("FF6B6B" if val < 0 else "FFFFFF"), size=10)
        else:
            c.value     = "N/A"
            c.font      = cell_font(color="888888", size=10)
            c.alignment = Alignment(horizontal="center")

    # Separator
    current_row += 1
    for ci in range(1, len(HEADERS) + 1):
        c = ws.cell(row=current_row, column=ci, value="")
        c.fill = PatternFill(start_color="2A475E", end_color="2A475E", fill_type="solid")

    # Rând TOTAL
    current_row += 1
    total_header_cell(current_row, 1, "📦 TOTAL PORTOFOLIU")
    for ci in range(2, 6):
        ws.cell(row=current_row, column=ci).fill   = TOTAL_FILL
        ws.cell(row=current_row, column=ci).border = THIN_BORDER
    ws.cell(row=current_row, column=3, value="Total acum ($):").font = cell_font(bold=True, color="FFFFFF")
    ws.cell(row=current_row, column=3).fill   = TOTAL_FILL
    ws.cell(row=current_row, column=3).border = THIN_BORDER
    total_value_cell(current_row, 4, total_sum, color="E67E22")
    total_pct_cell(current_row, 5, pct_g)

    col = 6
    for pk, _ in PERIOD_KEYS:
        pt           = period_totals.get(pk)
        pct_vs_now   = ((total_sum - pt) / pt * 100) if (pt and pt > 0) else None
        total_value_cell(current_row, col,     pt if pt is not None else "N/A", color="C6D4DF" if pt else "888888")
        total_pct_cell  (current_row, col + 1, pct_vs_now)
        col += 2
    ws.cell(row=current_row, column=col).fill   = TOTAL_FILL
    ws.cell(row=current_row, column=col).border = THIN_BORDER
    ws.row_dimensions[current_row].height = 20

    # Rând TAXĂ CSFloat
    current_row += 1
    for ci in range(1, len(HEADERS) + 1):
        c = ws.cell(row=current_row, column=ci)
        c.fill   = TAX_FILL
        c.border = THIN_BORDER
    ws.cell(row=current_row, column=3, value="Taxă CSFloat (2%):").font = cell_font(bold=True, color="FF9999")
    ws.cell(row=current_row, column=3).fill   = TAX_FILL
    ws.cell(row=current_row, column=3).border = THIN_BORDER
    c = ws.cell(row=current_row, column=4, value=round(total_sum * CSFLOAT_TAX, 2))
    c.number_format = "0.00"
    c.font  = cell_font(bold=True, color="FF9999")
    c.fill  = TAX_FILL
    c.border = THIN_BORDER

    col = 6
    for pk, _ in PERIOD_KEYS:
        pt      = period_totals.get(pk)
        tax_val = round(pt * CSFLOAT_TAX, 2) if pt else None
        c = ws.cell(row=current_row, column=col, value=tax_val if tax_val is not None else "N/A")
        c.font          = cell_font(bold=True, color="FF9999" if tax_val else "888888")
        c.fill          = TAX_FILL
        c.border        = THIN_BORDER
        c.number_format = "0.00"
        c.alignment     = Alignment(horizontal="center")
        ws.cell(row=current_row, column=col + 1).fill   = TAX_FILL
        ws.cell(row=current_row, column=col + 1).border = THIN_BORDER
        col += 2
    ws.cell(row=current_row, column=col).fill   = TAX_FILL
    ws.cell(row=current_row, column=col).border = THIN_BORDER

    # Rând NET
    current_row += 1
    for ci in range(1, len(HEADERS) + 1):
        c = ws.cell(row=current_row, column=ci)
        c.fill   = NET_FILL
        c.border = THIN_BORDER
    ws.cell(row=current_row, column=3, value="CE PRIMEȘTI (după taxă):").font = cell_font(bold=True, color="A4D007", size=11)
    ws.cell(row=current_row, column=3).fill   = NET_FILL
    ws.cell(row=current_row, column=3).border = THIN_BORDER
    c = ws.cell(row=current_row, column=4, value=round(total_sum * (1 - CSFLOAT_TAX), 2))
    c.number_format = "0.00"
    c.font  = cell_font(bold=True, color="A4D007", size=11)
    c.fill  = NET_FILL
    c.border = THIN_BORDER

    col = 6
    for pk, _ in PERIOD_KEYS:
        pt      = period_totals.get(pk)
        net_val = round(pt * (1 - CSFLOAT_TAX), 2) if pt else None
        c = ws.cell(row=current_row, column=col, value=net_val if net_val is not None else "N/A")
        c.font          = cell_font(bold=True, color="A4D007" if net_val else "888888", size=11)
        c.fill          = NET_FILL
        c.border        = THIN_BORDER
        c.number_format = "0.00"
        c.alignment     = Alignment(horizontal="center")
        ws.cell(row=current_row, column=col + 1).fill   = NET_FILL
        ws.cell(row=current_row, column=col + 1).border = THIN_BORDER
        col += 2
    ws.cell(row=current_row, column=col).fill   = NET_FILL
    ws.cell(row=current_row, column=col).border = THIN_BORDER
    ws.row_dimensions[current_row].height = 22

    # Notă
    current_row += 2
    note = ws.cell(row=current_row, column=1,
        value=f"ℹ️  Graficele se află în folderul '{CHARTS_DIR}/' — apasă pe numele unui item (portocaliu) pentru a deschide graficul în browser.")
    note.font = Font(name="Arial", italic=True, size=9, color="555555")

    wb.save(output)
    save_history(history)

    upload_all_to_gdrive(output, CHARTS_DIR)

    print("\n" + "=" * 60)
    print(f"✅ Salvat: {output}")
    print(f"📊 Grafice HTML: {len(chart_files)} fișiere în folderul '{CHARTS_DIR}/'")
    print(f"💶 Total: €{total_sum:.2f}  |  Net (după 2% taxă): €{total_sum * (1 - CSFLOAT_TAX):.2f}")
    if pct_g is not None:
        print(f"{'📈' if pct_g > 0 else '📉'} Global vs ultima rulare: {pct_g:+.2f}%")
    print(f"✅ OK: {ok}  ❌ Eșuat: {fail}")
    print("=" * 60)
    print(f"\n💡 TIP: Ține folderul '{CHARTS_DIR}/' în același loc cu '{output}'")
    print("   La click pe un item în Excel se deschide graficul în browser.")


# ── MAIN ────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("   CSFLOAT PRICE CALCULATOR – BUY NOW – GRAFICE HTML")
    print("=" * 60)

    items = [
        {'name': '★ Nomad Knife | Marble Fade (Factory New)', 'app_id': '730', 'quantity': 1},
        {'name': 'Zeus x27 | Tosai (Factory New)', 'app_id': '730', 'quantity': 80},
        {'name': 'AK-47 | Searing Rage (Factory New)', 'app_id': '730', 'quantity': 1},
        {'name': 'AK-47 | The Outsiders (Minimal Wear)', 'app_id': '730', 'quantity': 1},
        {'name': 'AK-47 | Aphrodite (Well-Worn)', 'app_id': '730', 'quantity': 1},
        {'name': 'AWP | Ice Coaled (Minimal Wear)', 'app_id': '730', 'quantity': 2},
		{'name': 'AWP | Ice Coaled (Factory New)', 'app_id': '730', 'quantity': 1},
		{'name': 'M4A4 | Turbine (Minimal Wear)', 'app_id': '730', 'quantity': 7},
        {'name': 'M4A1-S | Black Lotus (Factory New)', 'app_id': '730', 'quantity': 1},
        {'name': 'Desert Eagle | Mulberry (Factory New)', 'app_id': '730', 'quantity': 1},
        {'name': 'MAC-10 | Disco Tech (Minimal Wear)', 'app_id': '730', 'quantity': 1},
        {'name': 'P2000 | Acid Etched (Factory New)', 'app_id': '730', 'quantity': 1},
        {'name': 'P90 | Wave Breaker (Factory New)', 'app_id': '730', 'quantity': 1},
        {'name': 'FAMAS | 2A2F (Factory New)', 'app_id': '730', 'quantity': 1},
        {'name': 'AUG | Aristocrat (Factory New)', 'app_id': '730', 'quantity': 1},
        {'name': 'M4A4 | Tooth Fairy (Minimal Wear)', 'app_id': '730', 'quantity': 1},
        {'name': 'Glock-18 | Shinobu (Minimal Wear)', 'app_id': '730', 'quantity': 1},
        {'name': 'MP7 | Fade (Factory New)', 'app_id': '730', 'quantity': 1},
        {'name': 'MP9 | Mount Fuji (Minimal Wear)', 'app_id': '730', 'quantity': 1},
	{'name': 'P90 | Deathgaze (Minimal Wear)', 'app_id': '730', 'quantity': 1},
	{'name': 'Galil AR | Galigator (Factory New)', 'app_id': '730', 'quantity': 1},
	{'name': 'XM1014 | Teclu Burner (Factory New)', 'app_id': '730', 'quantity': 1},
	{'name': 'P250 | Kintsugi (Field-Tested)', 'app_id': '730', 'quantity': 1},
        {'name': 'Nova | Rising Sun (Factory New)', 'app_id': '730', 'quantity': 5},
        {'name': "'Two Times' McCoy | TACP Cavalry", 'app_id': '730', 'quantity': 1},
        {'name': 'Bloody Darryl The Strapped | The Professionals', 'app_id': '730', 'quantity': 1},
	{'name': '1st Lieutenant Farlow | SWAT', 'app_id': '730', 'quantity': 1},
	{'name': "'The Doctor' Romanov | Sabre", 'app_id': '730', 'quantity': 1},
        {'name': "Chef d'Escadron Rouchard | Gendarmerie Nationale", 'app_id': '730', 'quantity': 1},
        {'name': "Prof. Shahmat | Elite Crew", 'app_id': '730', 'quantity': 1},
        {'name': 'Clutch Case', 'app_id': '730', 'quantity': 50},
        {'name': 'Prisma 2 Case', 'app_id': '730', 'quantity': 50},
        {'name': 'Gallery Case', 'app_id': '730', 'quantity': 200},
        {'name': 'Danger Zone Case', 'app_id': '730', 'quantity': 5},
	{'name': 'Horizon Case', 'app_id': '730', 'quantity': 1},
	{'name': 'Chroma Case', 'app_id': '730', 'quantity': 1},
	{'name': 'Operation Breakout Weapon Case', 'app_id': '730', 'quantity': 10},
	{'name': 'Operation Vanguard Weapon Case', 'app_id': '730', 'quantity': 10},
        {'name': 'Fever Case', 'app_id': '730', 'quantity': 500},
        {'name': 'Fracture Case', 'app_id': '730', 'quantity': 150},
        {'name': 'Kilowatt Case', 'app_id': '730', 'quantity': 20},
        {'name': 'Recoil Case', 'app_id': '730', 'quantity': 150},
        {'name': 'Revolution Case', 'app_id': '730', 'quantity': 300},
        {'name': 'Snakebite Case', 'app_id': '730', 'quantity': 150},
	{'name': 'Dreams & Nightmares Case', 'app_id': '730', 'quantity': 50},
        {'name': 'Budapest 2025 Contenders Sticker Capsule', 'app_id': '730', 'quantity': 1000},
        {'name': 'Austin 2025 Contenders Sticker Capsule', 'app_id': '730', 'quantity': 100},
	{'name': 'Austin 2025 Legends Sticker Capsule', 'app_id': '730', 'quantity': 100},
        {'name': 'MP9 | Nexus (Field-Tested)', 'app_id': '730', 'quantity': 57},
        {'name': 'MAG-7 | Resupply (Field-Tested)', 'app_id': '730', 'quantity': 37},
        {'name': 'XM1014 | Mockingbird (Field-Tested)', 'app_id': '730', 'quantity': 5},
        {'name': '2021 Community Sticker Capsule', 'app_id': '730', 'quantity': 10},
	{'name': 'Community Sticker Capsule 1', 'app_id': '730', 'quantity': 5},
        {'name': 'Souvenir Charm | Austin 2025 Highlight | 1 Bullet', 'app_id': '730', 'quantity': 1},
        {'name': 'Sticker | torzsi (Gold) | Shanghai 2024', 'app_id': '730', 'quantity': 1},
        {'name': 'Sticker | iM (Holo) | Paris 2023', 'app_id': '730', 'quantity': 1},
        {'name': 'Sticker | iM (Gold) | Paris 2023', 'app_id': '730', 'quantity': 1},
        {'name': 'Sticker | TRAVIS (Holo) | Copenhagen 2024', 'app_id': '730', 'quantity': 2},
        {'name': 'Sticker | Natus Vincere (Holo) | Shanghai 2024', 'app_id': '730', 'quantity': 1},
        {'name': 'Sticker | Perfecto (Holo) | Paris 2023', 'app_id': '730', 'quantity': 1},
        {'name': 'Sticker | Queen Ava (Foil)', 'app_id': '730', 'quantity': 1},
        {'name': 'Sticker | b1t (Holo) | Paris 2023', 'app_id': '730', 'quantity': 1},
        {'name': 'Sticker | Imperial Esports (Glitter) | Copenhagen 2024', 'app_id': '730', 'quantity': 2},
        {'name': 'Sticker | Flex', 'app_id': '730', 'quantity': 2},
        {'name': 'Sticker | Ribbon Tie', 'app_id': '730', 'quantity': 1},
        {'name': 'Sticker | Complexity Gaming (Foil) | Austin 2025', 'app_id': '730', 'quantity': 3},
        {'name': 'Sticker | Quick Peek', 'app_id': '730', 'quantity': 100},
        {'name': 'Sticker | Taste Bud', 'app_id': '730', 'quantity': 20},
        {'name': 'Sticker | Overloaded (Glitter)', 'app_id': '730', 'quantity': 1},
        {'name': 'Sticker | Boom Blast (Glitter)', 'app_id': '730', 'quantity': 1},
        {'name': 'Sticker | Heroic (Gold) | 2020 RMR', 'app_id': '730', 'quantity': 1},
        {'name': 'Sticker | Fly High', 'app_id': '730', 'quantity': 15},
        {'name': 'Sticker | Hydro Stream', 'app_id': '730', 'quantity': 1},
        {'name': 'Sticker | iM (Glitter, Champion) | Copenhagen 2024', 'app_id': '730', 'quantity': 20},
        {'name': 'Sticker | Scorch Loop (Reverse)', 'app_id': '730', 'quantity': 124},
        {'name': 'Sticker | XD', 'app_id': '730', 'quantity': 20},
        {'name': 'Sticker | Glare', 'app_id': '730', 'quantity': 20},
        {'name': 'Sticker | From The Deep', 'app_id': '730', 'quantity': 20},
        {'name': 'Sticker | From The Deep (Glitter)', 'app_id': '730', 'quantity': 20},
        {'name': 'Sticker | Chompers', 'app_id': '730', 'quantity': 1},
        {'name': 'Sticker | Heroic (Holo) | 2020 RMR', 'app_id': '730', 'quantity': 20},
        {'name': 'Sticker | Ninjas in Pyjamas (Holo) | 2020 RMR', 'app_id': '730', 'quantity': 25},
        {'name': 'Sticker | Scorch Loop', 'app_id': '730', 'quantity': 25},
        {'name': 'Sticker | Boom Epicenter (Glitter)', 'app_id': '730', 'quantity': 25},
        {'name': 'Sticker | Bolt Energy', 'app_id': '730', 'quantity': 50},
        {'name': 'Sticker | Hydro Stream', 'app_id': '730', 'quantity': 100},
        {'name': 'Sticker | Bomb Planted (Holo)', 'app_id': '730', 'quantity': 4},
        {'name': 'Sticker | M80 (Foil) | Austin 2025', 'app_id': '730', 'quantity': 10},
        {'name': 'Sticker | FaZe Clan (Foil) | Austin 2025', 'app_id': '730', 'quantity': 20},
        {'name': 'Sticker | iM (Foil) | Austin 2025', 'app_id': '730', 'quantity': 100},
        {'name': 'Sticker | NiKo | Paris 2023', 'app_id': '730', 'quantity': 200},
        {'name': 'Sticker | RED Canids (Gold) | Budapest 2025', 'app_id': '730', 'quantity': 2},
        {'name': 'Sticker | fnatic (Gold) | Budapest 2025', 'app_id': '730', 'quantity': 1},
        {'name': 'Sticker | iM (Glitter) | Copenhagen 2024', 'app_id': '730', 'quantity': 50}
    ]

    global USD_TO_EUR
    USD_TO_EUR = get_usd_to_eur()

    print(f"\n📋 {len(items)} iteme de procesat\n")
    create_excel(items, "csfloat_items.xlsx")


if __name__ == "__main__":
    main()
